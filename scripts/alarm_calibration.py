"""Historical calibration of the Lens 2 composite alarm level.

Usage:
    python scripts/alarm_calibration.py

Reconstructs the froth composite monthly with point-in-time discipline and
evaluates candidate alarm levels (50 / 62.5 / 75 / 87.5 percent of gauges
triggered) against S&P 500 forward returns over 1, 3, 6, and 12 months.
Findings are filed in reviews/ and the studies ledger; the alarm level
itself remains ZH's decision (SPEC.md section 3).

Anti-look-ahead construction:
- Percentile triggers (AAII decile, P/E, Rule-of-20 tail, NFCI loose tail,
  Michigan sentiment) use EXPANDING windows: the line at month t uses only
  data up to t. Absolute triggers (NAAIM 90, value-growth +10 pp, the 20 in
  Rule-of-20) are point-in-time by construction.
- The IPO gauge is excluded (history starts 2016 — too short to
  reconstruct), so the composite here is 5 gauges from 1990 and 7 gauges
  from 2007.

Known residual biases (documented, not fixable free): current-vintage
histories (NFCI re-estimates weekly; Michigan and CPI are near-unrevised)
and month-end availability approximations for monthly prints (up to a few
weeks optimistic). Episode counts are small; per-episode outcomes are
reported alongside pooled statistics for exactly that reason.

Date handling: Python datetime months are 1-indexed; month arithmetic uses
dateutil.relativedelta only. The AAII and NAAIM workbooks and the multpl
table are used in-memory and not redistributed.
"""
from __future__ import annotations

import io
import json
import re
from datetime import date
from pathlib import Path

import openpyxl
from dateutil.relativedelta import relativedelta

from sources import sentiment
from sources.fred import fetch_series
from sources.prices import fetch_yahoo_daily
from sources.scrape import fetch_bytes, fetch_text
from util import ROOT, clean_series

CANDIDATE_LEVELS = (50.0, 62.5, 75.0, 87.5)
HORIZONS_MONTHS = (1, 3, 6, 12)
PANELS = (("A: 5-gauge, 1990-01 onwards", date(1990, 1, 31), 5),
          ("B: 7-gauge, 2007-01 onwards", date(2007, 1, 31), 7))


def expanding_percentile_rank(sorted_history_getter, value):
    history = sorted_history_getter
    if not history:
        return None
    return 100.0 * sum(1 for v in history if v <= value) / len(history)


def last_at_or_before(dates: list[str], values: list[float], iso: str):
    """Latest observation at or before iso date; None when none exists."""
    result = None
    for d, v in zip(dates, values):
        if d <= iso:
            result = v
        else:
            break
    return result


def series_upto(dates: list[str], values: list[float], iso: str) -> list[float]:
    return [v for d, v in zip(dates, values) if d <= iso]


def fetch_naaim_history() -> tuple[list[str], list[float]]:
    page = fetch_text(sentiment.NAAIM_URL)
    link = re.search(r'href="(https://naaim\.org/[^"]*USE_Data[^"]*\.xlsx)"', page)
    if not link:
        raise RuntimeError("NAAIM history workbook link not found")
    book = openpyxl.load_workbook(io.BytesIO(fetch_bytes(link.group(1))), read_only=True)
    sheet = book[book.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        rows.append((row[0].date().isoformat(), float(row[1])))
    rows.sort()
    return [d for d, _ in rows], [v for _, v in rows]


def fetch_aaii_history() -> tuple[list[str], list[float]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=fetch_bytes(sentiment.AAII_XLS_URL))
    sheet = book.sheet_by_name("SENTIMENT")
    rows = []
    for r in range(sheet.nrows):
        serial, spread = sheet.cell_value(r, 0), sheet.cell_value(r, 6)
        if isinstance(serial, float) and isinstance(spread, float):
            day = xlrd.xldate_as_datetime(serial, book.datemode).date()
            rows.append((day.isoformat(), spread * 100.0))
    rows.sort()
    return [d for d, _ in rows], [v for _, v in rows]


def month_ends_from_daily(dates: list[str]) -> list[str]:
    """Last trading day per calendar month (ISO strings, sorted input)."""
    ends = []
    for i, d in enumerate(dates):
        if i + 1 == len(dates) or dates[i + 1][:7] != d[:7]:
            ends.append(d)
    return ends


def max_within(values: list[float]) -> float:
    return max(values) if values else float("nan")


def build_composite_grid() -> list[dict]:
    """Monthly point-in-time Lens 2 composite grid.

    Shared machinery: the filed calibration (main below) and the phase 5
    forward-return study both consume this. Returns one dict per month-end
    with iso date, gauge count, composite share, and per-gauge statuses.
    """
    print("Fetching histories (all in-memory)...")
    aaii_d, aaii_v = fetch_aaii_history()
    naaim_d, naaim_v = fetch_naaim_history()
    pe_d, pe_v, _ = sentiment.fetch_multpl_pe_history()
    umc_d, umc_v = clean_series(*fetch_series("UMCSENT"))
    cpi_d, cpi_v = clean_series(*fetch_series("CPIAUCSL"))
    nfci_d, nfci_v = clean_series(*fetch_series("NFCI"))
    rpg_d, rpg_v = fetch_yahoo_daily("RPG", period1=0)
    rpv_d, rpv_v = fetch_yahoo_daily("RPV", period1=0)

    gspc = json.loads((ROOT / "data" / "history" / "gspc.json").read_text(encoding="utf-8"))
    spx_ends = month_ends_from_daily(gspc["dates"])
    spx_close = dict(zip(gspc["dates"], gspc["values"]))

    cpi_by = dict(zip(cpi_d, cpi_v))

    # Precompute CPI year-on-year per reference month (one pass), then the
    # Rule-of-20 sum series aligned to multpl P/E months from 1948 — the
    # expanding percentile below only slices this precomputed series.
    cpi_yoy_by_month: dict[str, float] = {}
    for month in cpi_d:
        base = (date.fromisoformat(month) - relativedelta(years=1)).isoformat()
        if base in cpi_by:
            cpi_yoy_by_month[month] = (cpi_by[month] / cpi_by[base] - 1.0) * 100.0

    def cpi_yoy_at(iso: str):
        month = next((d for d in reversed(cpi_d) if d <= iso), None)
        return cpi_yoy_by_month.get(month) if month else None

    rule20_dates: list[str] = []
    rule20_sums: list[float] = []
    for d, v in zip(pe_d, pe_v):
        if d >= "1948-01-01" and d[:8] + "01" in cpi_yoy_by_month:
            rule20_dates.append(d)
            rule20_sums.append(v + cpi_yoy_by_month[d[:8] + "01"])

    rpg_by, rpv_by = dict(zip(rpg_d, rpg_v)), dict(zip(rpv_d, rpv_v))
    common_px = sorted(set(rpg_d) & set(rpv_d))

    def value_growth_spread_at(iso: str):
        window = [d for d in common_px if d <= iso]
        if len(window) < 127:
            return None
        last, past = window[-1], window[-127]
        return ((rpg_by[last] / rpg_by[past]) - (rpv_by[last] / rpv_by[past])) * 100.0

    # Monthly grid with point-in-time gauge statuses.
    grid = []
    for iso in spx_ends:
        t = date.fromisoformat(iso)
        if t < date(1988, 1, 1):
            continue
        statuses = {}

        aaii_hist = series_upto(aaii_d, aaii_v, iso)
        aaii_now = last_at_or_before(aaii_d, aaii_v, iso)
        if aaii_now is not None and len(aaii_hist) >= 156:  # 3-year burn-in
            line = sorted(aaii_hist)[int(0.9 * (len(aaii_hist) - 1))]
            statuses["aaii"] = aaii_now >= line

        pe_hist = series_upto(pe_d, pe_v, iso)
        pe_now = last_at_or_before(pe_d, pe_v, iso)
        if pe_now is not None and pe_hist:
            statuses["pe"] = expanding_percentile_rank(pe_hist, pe_now) >= 90.0

        yoy = cpi_yoy_at(iso)
        if pe_now is not None and yoy is not None:
            sums_hist = series_upto(rule20_dates, rule20_sums, iso)
            total = pe_now + yoy
            statuses["rule20"] = bool(
                sums_hist
                and total > 20.0
                and expanding_percentile_rank(sums_hist, total) >= 80.0
            )

        nfci_hist = series_upto(nfci_d, nfci_v, iso)
        nfci_now = last_at_or_before(nfci_d, nfci_v, iso)
        if nfci_now is not None and nfci_hist:
            statuses["nfci"] = expanding_percentile_rank(nfci_hist, nfci_now) <= 20.0

        umc_hist = series_upto(umc_d, umc_v, iso)
        umc_now = last_at_or_before(umc_d, umc_v, iso)
        if umc_now is not None and umc_hist:
            statuses["umcsent"] = expanding_percentile_rank(umc_hist, umc_now) >= 75.0

        naaim_now = last_at_or_before(naaim_d, naaim_v, iso)
        if naaim_now is not None:
            statuses["naaim"] = naaim_now >= 90.0

        vvg = value_growth_spread_at(iso)
        if vvg is not None:
            statuses["vvg"] = vvg >= 10.0

        grid.append({
            "iso": iso,
            "n": len(statuses),
            "share": 100.0 * sum(statuses.values()) / len(statuses) if statuses else None,
            "statuses": statuses,
        })

    return grid


def load_spx_month_ends() -> tuple[list[str], dict[str, float]]:
    """Month-end ISO dates and close lookup from the stored S&P history."""
    gspc = json.loads((ROOT / "data" / "history" / "gspc.json").read_text(encoding="utf-8"))
    return month_ends_from_daily(gspc["dates"]), dict(zip(gspc["dates"], gspc["values"]))


def main() -> int:
    grid = build_composite_grid()
    _, spx_close = load_spx_month_ends()
    spx_series = [(g["iso"], spx_close[g["iso"]]) for g in grid]

    def forward_return(index: int, months: int):
        if index + months < len(spx_series):
            return (spx_series[index + months][1] / spx_series[index][1] - 1.0) * 100.0
        return None

    def worst_within_12(index: int):
        base = spx_series[index][1]
        outs = [spx_series[index + k][1] / base - 1.0
                for k in range(1, 13) if index + k < len(spx_series)]
        return min(outs) * 100.0 if outs else None

    for title, start, min_gauges in PANELS:
        rows = [(i, g) for i, g in enumerate(grid)
                if date.fromisoformat(g["iso"]) >= start and g["n"] >= min_gauges]
        print(f"\n=== Panel {title} ({len(rows)} months, "
              f"{rows[0][1]['iso']} to {rows[-1][1]['iso']}) ===")

        uncond = {h: [r for i, g in rows if (r := forward_return(i, h)) is not None]
                  for h in HORIZONS_MONTHS}
        base_line = "  ".join(
            f"{h:>2}m med {sorted(uncond[h])[len(uncond[h]) // 2]:+5.1f}% hit {100 * sum(1 for r in uncond[h] if r > 0) / len(uncond[h]):3.0f}%"
            for h in HORIZONS_MONTHS)
        print(f"Unconditional ({len(uncond[12])} obs at 12m): {base_line}")

        for level in CANDIDATE_LEVELS:
            active = [(i, g) for i, g in rows if g["share"] is not None and g["share"] >= level]
            if not active:
                print(f"\n>= {level:.1f}%: never active")
                continue
            print(f"\n>= {level:.1f}%: {len(active)} active months")
            for h in HORIZONS_MONTHS:
                fwd = [r for i, g in active if (r := forward_return(i, h)) is not None]
                if not fwd:
                    continue
                med = sorted(fwd)[len(fwd) // 2]
                mean = sum(fwd) / len(fwd)
                hit = 100.0 * sum(1 for r in fwd if r > 0) / len(fwd)
                print(f"   {h:>2}m: n={len(fwd):3d} mean {mean:+6.2f}% med {med:+6.2f}% hit {hit:3.0f}%")
            # Episodes: contiguous active runs on the monthly grid.
            episodes = []
            run = []
            active_isos = {g["iso"] for i, g in active}
            for i, g in rows:
                if g["iso"] in active_isos:
                    run.append((i, g))
                elif run:
                    episodes.append(run)
                    run = []
            if run:
                episodes.append(run)
            print(f"   episodes ({len(episodes)}):")
            for ep in episodes:
                i0, g0 = ep[0]
                i1, g1 = ep[-1]
                peak = max(g["share"] for _, g in ep)
                f12 = forward_return(i0, 12)
                worst = worst_within_12(i0)
                f12_text = f"{f12:+6.1f}%" if f12 is not None else "  n/a "
                worst_text = f"{worst:+6.1f}%" if worst is not None else "  n/a "
                print(f"     {g0['iso']} to {g1['iso']} ({len(ep):2d}m, peak {peak:3.0f}%) "
                      f"fwd12 from onset {f12_text}, worst point in 12m {worst_text}")
    return 0


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    raise SystemExit(main())
